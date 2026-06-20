#!/usr/bin/env python3
"""CLI that exports competitor matrix data from JSONL to a multi-sheet XLSX workbook.

Reads a JSONL file containing product, price, and claims data and produces a
formatted Excel workbook with three sheets:

- **Products** — product identity: name, brand, manufacturer, dosage form,
  dose, active ingredients, package quantity, version identifier.
- **Prices** — price observations: price amount, currency, price type,
  unit price, daily cost, collection date, seller type, source platform.
- **Claims** — commercial claims: claim text, claim type, support level
  (when verified), evidence references, source platform, extract date.

Empty records with no relevant data for a sheet are still written (one
header-only row) to keep sheets consistent.

Usage
-----
    python scripts/export-xlsx.py input.jsonl --out output.xlsx

Input
-----
JSONL file where each record is a dict that may contain product-identity
fields (``standard_name``, ``brand``, ``dosage_form``, etc.),
``price_observations`` (list of dicts), and/or ``claims`` (list of dicts).

Exit codes
----------
0   Export successful — XLSX written.
2   File not found or missing arguments.
3   Corrupted JSONL, openpyxl import failure, or write error.

The input file is never modified.
"""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path
from typing import Any, Dict, List

# ---------------------------------------------------------------------------
# Project root on sys.path for ``scripts._common`` import.
# ---------------------------------------------------------------------------
_T = Path(__file__).resolve().parent.parent
if str(_T) not in sys.path:
    sys.path.insert(0, str(_T))

from scripts._common import read_jsonl, log_info, log_error

# ---------------------------------------------------------------------------
# Lazy import of openpyxl
# ---------------------------------------------------------------------------

_OPENPYXL_LOADED = False


def _import_openpyxl() -> None:
    """Import ``openpyxl`` once, failing with exit code 3 if not installed."""
    global _OPENPYXL_LOADED
    if _OPENPYXL_LOADED:
        return
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        log_error(
            "openpyxl is required for XLSX export — "
            "install with: pip install openpyxl"
        )
        sys.exit(3)
    _OPENPYXL_LOADED = True


# ---------------------------------------------------------------------------
# Column definitions
# ---------------------------------------------------------------------------

PRODUCT_COLUMNS = [
    "product",
    "brand",
    "manufacturer",
    "dosage_form",
    "dose",
    "active_ingredients",
    "package_quantity",
    "package_description",
    "version_id",
    "nmpa_approval",
    "otc_rx",
    "data_date",
    "notes",
]

PRICE_COLUMNS = [
    "product",
    "brand",
    "price",
    "currency",
    "price_type",
    "unit_price",
    "unit_price_currency",
    "unit_price_unit",
    "daily_cost",
    "package_quantity",
    "seller_type",
    "source_platform",
    "collection_date",
    "url",
]

CLAIMS_COLUMNS = [
    "product",
    "brand",
    "claim_text",
    "claim_type",
    "support_level",
    "evidence_ids",
    "verdict",
    "source_platform",
    "extract_date",
    "notes",
]


# ---------------------------------------------------------------------------
# Row extraction
# ---------------------------------------------------------------------------


def _fmt_cell(val: Any) -> Any:
    """Format a value for XLSX cell output.

    Parameters
    ----------
    val : Any
        Raw value from a JSONL record.

    Returns
    -------
    str or the original value
        Lists are joined with ``"; "``; other values are returned as-is.
    """
    if isinstance(val, list):
        return "; ".join(str(v) for v in val)
    if val is None:
        return ""
    return val


def _extract_product_rows(
    records: List[Dict[str, Any]],
) -> List[Dict[str, Any]]:
    """Extract product-identity rows from all records.

    Each input record produces one output row.  Missing fields are set to the
    empty string.

    Parameters
    ----------
    records : list[dict]
        Input JSONL records.

    Returns
    -------
    list[dict]
        One dict per input record, keyed by ``PRODUCT_COLUMNS``.
    """
    rows: List[Dict[str, Any]] = []
    for rec in records:
        ingredients = rec.get("active_ingredients", [])
        if ingredients and isinstance(ingredients, list):
            ing_texts = []
            for ing in ingredients:
                if isinstance(ing, dict):
                    name = ing.get("name", "")
                    dose = ing.get("per_unit_dose", "")
                    unit = ing.get("per_unit_dose_unit", "")
                    ing_texts.append(f"{name} {dose} {unit}".strip())
                else:
                    ing_texts.append(str(ing))
            ing_str = "; ".join(ing_texts)
        else:
            ing_str = ""

        dose_raw = ""
        if ingredients and isinstance(ingredients, list) and len(ingredients) > 0:
            first = ingredients[0]
            d = first.get("per_unit_dose", "")
            u = first.get("per_unit_dose_unit", "")
            if d and d not in ("", "unknown"):
                dose_raw = f"{d} {u}".strip()

        row = {
            "product": rec.get("standard_name") or rec.get("generic_name", ""),
            "brand": rec.get("brand", ""),
            "manufacturer": rec.get("manufacturer", ""),
            "dosage_form": rec.get("dosage_form", ""),
            "dose": dose_raw,
            "active_ingredients": ing_str,
            "package_quantity": _fmt_cell(rec.get("package_quantity", "")),
            "package_description": rec.get("package_description", ""),
            "version_id": rec.get("product_version_id", ""),
            "nmpa_approval": rec.get("nmpa_approval_number", ""),
            "otc_rx": rec.get("otc_rx", ""),
            "data_date": rec.get("collection_date", ""),
            "notes": rec.get("data_quality_notes", ""),
        }
        rows.append(row)
    return rows


def _extract_price_rows(
    records: List[Dict[str, Any]],
) -> List[Dict[str, Any]]:
    """Extract price-observation rows from all records.

    Each ``price_observations`` entry produces one output row.  If a record
    has no price observations, it is skipped for this sheet.

    Parameters
    ----------
    records : list[dict]
        Input JSONL records.

    Returns
    -------
    list[dict]
        One dict per price observation, keyed by ``PRICE_COLUMNS``.
    """
    rows: List[Dict[str, Any]] = []
    for rec in records:
        product_name = rec.get("standard_name") or rec.get("generic_name", "")
        brand = rec.get("brand", "")
        obs_list = rec.get("price_observations", [])
        if not isinstance(obs_list, list):
            continue
        for obs in obs_list:
            if not isinstance(obs, dict):
                continue
            row = {
                "product": product_name,
                "brand": brand,
                "price": obs.get("price", ""),
                "currency": obs.get("currency", ""),
                "price_type": obs.get("price_type", ""),
                "unit_price": obs.get("unit_price", ""),
                "unit_price_currency": obs.get("unit_price_currency", ""),
                "unit_price_unit": obs.get("unit_price_unit", ""),
                "daily_cost": obs.get("daily_cost", ""),
                "package_quantity": _fmt_cell(obs.get("package_quantity", "")),
                "seller_type": obs.get("seller_type", ""),
                "source_platform": obs.get("source_platform", obs.get("source_id", "")),
                "collection_date": obs.get("collection_date", ""),
                "url": obs.get("url", ""),
            }
            rows.append(row)
    return rows


def _extract_claims_rows(
    records: List[Dict[str, Any]],
) -> List[Dict[str, Any]]:
    """Extract commercial-claim rows from all records.

    Each entry in a record's ``claims`` array (or a record-level
    ``claims_summary``) produces one output row.  If a record has no claims,
    it is skipped.

    Parameters
    ----------
    records : list[dict]
        Input JSONL records.

    Returns
    -------
    list[dict]
        One dict per claim, keyed by ``CLAIMS_COLUMNS``.
    """
    rows: List[Dict[str, Any]] = []
    for rec in records:
        product_name = rec.get("standard_name") or rec.get("generic_name", "")
        brand = rec.get("brand", "")

        claims = rec.get("claims", rec.get("commercial_claims", []))
        if isinstance(claims, list) and len(claims) > 0:
            for claim in claims:
                if not isinstance(claim, dict):
                    continue
                row = {
                    "product": product_name,
                    "brand": brand,
                    "claim_text": claim.get("claim_text", claim.get("text", "")),
                    "claim_type": claim.get("claim_type", claim.get("type", "")),
                    "support_level": claim.get("support_level", claim.get(
                        "claim_support_level", ""
                    )),
                    "evidence_ids": _fmt_cell(
                        claim.get("evidence_ids", claim.get("supporting_evidence", []))
                    ),
                    "verdict": claim.get("verdict", claim.get("review_verdict", "")),
                    "source_platform": claim.get(
                        "source_platform", claim.get("source_id", "")
                    ),
                    "extract_date": claim.get("extract_date", claim.get(
                        "collection_date", ""
                    )),
                    "notes": claim.get("reviewer_notes", ""),
                }
                rows.append(row)

        # Fall back to claims_summary if no list
        elif rec.get("claims_summary"):
            row = {
                "product": product_name,
                "brand": brand,
                "claim_text": rec["claims_summary"],
                "claim_type": "summary",
                "support_level": "",
                "evidence_ids": "",
                "verdict": "",
                "source_platform": rec.get("source_id", ""),
                "extract_date": rec.get("collection_date", ""),
                "notes": "",
            }
            rows.append(row)

    return rows


# ---------------------------------------------------------------------------
# XLSX generation
# ---------------------------------------------------------------------------


def _write_sheet(ws: Any, columns: List[str], rows: List[Dict[str, Any]]) -> None:
    """Populate an openpyxl worksheet with a header row and data rows.

    Parameters
    ----------
    ws : openpyxl.worksheet.worksheet.Worksheet
        Target worksheet.
    columns : list[str]
        Ordered column names.
    rows : list[dict]
        Data rows keyed by column names.
    """
    from openpyxl.utils import get_column_letter

    # Header
    ws.append(columns)

    # Data rows
    for row in rows:
        ws.append([_fmt_cell(row.get(col, "")) for col in columns])

    # Auto-adjust column widths (capped at 40)
    for col_idx, col_name in enumerate(columns, start=1):
        max_len = len(col_name)
        for row_idx in range(2, min(len(rows) + 2, 1002)):
            cell_val = ws.cell(row=row_idx, column=col_idx).value
            if cell_val is not None:
                max_len = max(max_len, len(str(cell_val)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(
            max_len + 2, 40
        )

    # Freeze top row
    ws.freeze_panes = "A2"


def _write_xlsx(
    path: Path,
    product_rows: List[Dict[str, Any]],
    price_rows: List[Dict[str, Any]],
    claims_rows: List[Dict[str, Any]],
) -> None:
    """Create a multi-sheet XLSX workbook and write to *path*.

    Parameters
    ----------
    path : Path
        Output ``.xlsx`` file path.
    product_rows : list[dict]
        Product-identity rows.
    price_rows : list[dict]
        Price observation rows.
    claims_rows : list[dict]
        Commercial claims rows.

    Raises
    ------
    SystemExit
        Exit 3 on openpyxl import failure.
    """
    _import_openpyxl()
    from openpyxl import Workbook

    wb = Workbook()

    # -- Products sheet --
    ws_products = wb.active
    ws_products.title = "Products"
    _write_sheet(ws_products, PRODUCT_COLUMNS, product_rows)

    # -- Prices sheet --
    ws_prices = wb.create_sheet("Prices")
    _write_sheet(ws_prices, PRICE_COLUMNS, price_rows)

    # -- Claims sheet --
    ws_claims = wb.create_sheet("Claims")
    _write_sheet(ws_claims, CLAIMS_COLUMNS, claims_rows)

    wb.save(str(path))


# ---------------------------------------------------------------------------
# CLI entry point
# ---------------------------------------------------------------------------


def main() -> int:
    """Parse arguments, read input, extract rows, write XLSX, return exit code."""
    parser = argparse.ArgumentParser(
        description="Export competitor matrix data from JSONL to XLSX.",
    )
    parser.add_argument(
        "input",
        type=str,
        help="Path to a .jsonl file containing product/price/claims data",
    )
    parser.add_argument(
        "--out",
        type=str,
        required=True,
        help="Output .xlsx file path",
    )
    args = parser.parse_args()

    input_path = Path(args.input)
    output_path = Path(args.out)

    # -- Validate input --
    if not input_path.exists():
        log_error(f"File not found: {input_path}")
        return 2

    # -- Read input --
    try:
        records = read_jsonl(input_path)
    except json.JSONDecodeError as exc:
        log_error(f"Corrupted JSONL at {input_path}: {exc}")
        return 3

    log_info(f"Read {len(records)} record(s) from {input_path}")

    # -- Extract rows --
    product_rows = _extract_product_rows(records)
    price_rows = _extract_price_rows(records)
    claims_rows = _extract_claims_rows(records)

    log_info(
        f"Extracted {len(product_rows)} product(s), "
        f"{len(price_rows)} price observation(s), "
        f"{len(claims_rows)} claim(s)"
    )

    # -- Write XLSX --
    try:
        _write_xlsx(output_path, product_rows, price_rows, claims_rows)
    except Exception as exc:
        log_error(f"XLSX export failed: {exc}")
        return 3

    log_info(f"XLSX exported: {output_path} ({len(product_rows)}/{len(price_rows)}/{len(claims_rows)} rows)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
