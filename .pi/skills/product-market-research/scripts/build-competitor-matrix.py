#!/usr/bin/env python3
"""CLI that reads merged product records and exports a competitor matrix CSV/XLSX.

Transforms and calculates only — no strategic conclusions are generated.

Usage
-----
    python scripts/build-competitor-matrix.py input.jsonl --out output.csv [--format csv]
    python scripts/build-competitor-matrix.py input.jsonl --out output.xlsx --format xlsx

Input
-----
JSONL file with merged product records. Each record should contain product-level
fields (matching ``product.schema.json``) and may also contain:

    price_observations : list[dict]
        Each dict with ``price``, ``price_type``, ``unit_price``, ``daily_cost``.
    channels : list[str]
        Platform codes where the product was observed.
    claims_summary : str
        Commercial claims summary.
    main_limitations : str
        Free-text limitations for this entry.
    collection_date : str
        ISO 8601 date of data collection.

Output columns (in order)
-------------------------
product, brand, version, dosage_form, dose, spec, price, price_type,
normalized_price, daily_cost, claims, channel, data_date, data_quality,
main_limitations

Exit codes
----------
0   Success — output written.
1   Required column(s) cannot be populated (product name, brand, or dosage_form
    missing from one or more records).
2   File not found, missing arguments, or unsupported format.
3   Corrupted JSONL or schema-loading error.

The input file is never modified.
"""

from __future__ import annotations

import argparse
import csv
import json
import sys
import warnings
from pathlib import Path
from typing import Any, Dict, List, Optional

# ---------------------------------------------------------------------------
# Ensure the project root is on sys.path so ``scripts._common`` is importable
# when run as ``python scripts/build-competitor-matrix.py``.
# ---------------------------------------------------------------------------
_PROJECT_ROOT = Path(__file__).resolve().parent.parent
if str(_PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(_PROJECT_ROOT))

from scripts._common import read_jsonl, log_info, log_error

try:
    with warnings.catch_warnings():
        warnings.filterwarnings("ignore", category=DeprecationWarning)
        from jsonschema import Draft7Validator, RefResolver
except ImportError:
    log_error("jsonschema is required — install with: pip install jsonschema")
    sys.exit(3)

# ---------------------------------------------------------------------------
# Output column definitions
# ---------------------------------------------------------------------------

#: Ordered list of output column names.
OUTPUT_COLUMNS = [
    "product",
    "brand",
    "version",
    "dosage_form",
    "dose",
    "spec",
    "price",
    "price_type",
    "normalized_price",
    "daily_cost",
    "claims",
    "channel",
    "data_date",
    "data_quality",
    "main_limitations",
]


def _extract_row(record: Dict[str, Any]) -> Dict[str, Any]:
    """Map a merged product record to the 15 output columns.

    Parameters
    ----------
    record : dict
        One merged product record from the input JSONL.

    Returns
    -------
    dict
        Row mapped to ``OUTPUT_COLUMNS`` keys.
    """
    # -- Product name ---------------------------------------------------------
    product = record.get("standard_name") or record.get("generic_name") or ""

    # -- Brand ----------------------------------------------------------------
    brand = record.get("brand") or ""

    # -- Version --------------------------------------------------------------
    version = record.get("record_version") or record.get("product_version_id") or ""
    if isinstance(version, int):
        version = str(version)

    # -- Dosage form ----------------------------------------------------------
    dosage_form = record.get("dosage_form") or ""

    # -- Dose (first active ingredient per-unit dose) -------------------------
    ingredients = record.get("active_ingredients", [])
    if ingredients and isinstance(ingredients, list) and len(ingredients) > 0:
        dose_raw = ingredients[0].get("per_unit_dose", "")
        dose_unit = ingredients[0].get("per_unit_dose_unit", "")
        if dose_raw and dose_raw not in ("", "unknown"):
            dose = f"{dose_raw} {dose_unit}".strip()
        else:
            dose = dose_raw or "unknown"
    else:
        dose = "unknown"

    # -- Spec (package quantity or description) --------------------------------
    spec = record.get("package_quantity") or record.get("package_description") or ""
    if isinstance(spec, (int, float)):
        spec = str(spec)

    # -- Price observations ---------------------------------------------------
    norm = record.get("normalized", {}) or {}
    price_obs = record.get("price_observations", [])

    if price_obs and isinstance(price_obs, list) and len(price_obs) > 0:
        first = price_obs[0]
        price = first.get("price", "")
        price_type = first.get("price_type", "")
    else:
        price = ""
        price_type = ""

    # Normalized price: prefer record['normalized']['unit_price'], then
    # price_observations[0]['normalized']['unit_price'], then
    # compute from price_observations price/package_quantity, then
    # fall back to price_obs unit_price.
    np_val = norm.get("unit_price")
    if np_val is not None:
        normalized_price = np_val
    elif (
        price_obs
        and isinstance(price_obs, list)
        and len(price_obs) > 0
    ):
        first = price_obs[0]
        obs_norm = first.get("normalized", {}) or {}
        obs_np = obs_norm.get("unit_price")
        if obs_np is not None:
            normalized_price = obs_np
        else:
            p = first.get("price")
            pkq = first.get("package_quantity")
            if isinstance(p, (int, float)) and isinstance(pkq, (int, float)) and pkq > 0:
                normalized_price = p / pkq
            else:
                normalized_price = first.get("unit_price", "")
    else:
        normalized_price = ""

    # Daily cost: prefer record['normalized']['daily_cost'], then
    # record['daily_cost'], then price_observations[0]['normalized']['daily_cost'],
    # then price_observations[0]['daily_cost'].
    dc_val = norm.get("daily_cost")
    if dc_val is not None:
        daily_cost = dc_val
    else:
        dc_val = record.get("daily_cost")
        if dc_val is not None:
            daily_cost = dc_val
        elif (
            price_obs
            and isinstance(price_obs, list)
            and len(price_obs) > 0
        ):
            first = price_obs[0]
            obs_norm = first.get("normalized", {}) or {}
            obs_dc = obs_norm.get("daily_cost")
            if obs_dc is not None:
                daily_cost = obs_dc
            else:
                daily_cost = first.get("daily_cost", "")
        else:
            daily_cost = ""

    # -- Claims ---------------------------------------------------------------
    claims = record.get("claims_summary") or "not_extracted"

    # -- Channel --------------------------------------------------------------
    # Prefer dedicated channels array, fall back to source_ids.
    channels = record.get("channels", [])
    if not channels:
        channels = record.get("source_ids", [])
    if isinstance(channels, list):
        channel = ",".join(str(c) for c in channels)
    else:
        channel = str(channels) if channels else ""

    # -- Data date ------------------------------------------------------------
    data_date = record.get("collection_date") or ""
    if not data_date and price_obs and isinstance(price_obs, list) and len(price_obs) > 0:
        data_date = price_obs[0].get("collection_date", "")

    # -- Data quality ---------------------------------------------------------
    data_quality = record.get("data_quality_score") or "unknown"

    # -- Main limitations -----------------------------------------------------
    main_limitations = record.get("main_limitations") or "none_recorded"

    return {
        "product": product,
        "brand": brand,
        "version": version,
        "dosage_form": dosage_form,
        "dose": dose,
        "spec": spec,
        "price": price,
        "price_type": price_type,
        "normalized_price": normalized_price,
        "daily_cost": daily_cost,
        "claims": claims,
        "channel": channel,
        "data_date": data_date,
        "data_quality": data_quality,
        "main_limitations": main_limitations,
    }


def _check_required_columns(rows: List[Dict[str, Any]]) -> bool:
    """Check that essential columns (product, brand, dosage_form) are populated.

    Parameters
    ----------
    rows : list[dict]
        Extracted rows.

    Returns
    -------
    bool
        ``True`` if all rows have product, brand, and dosage_form populated.
    """
    all_ok = True
    for idx, row in enumerate(rows):
        line = idx + 1
        missing = []
        if not row.get("product"):
            missing.append("product")
        if not row.get("brand"):
            missing.append("brand")
        if not row.get("dosage_form"):
            missing.append("dosage_form")
        if missing:
            log_error(
                f"Record {line}: required column(s) cannot be populated: "
                f"{', '.join(missing)}"
            )
            all_ok = False
    return all_ok


# ---------------------------------------------------------------------------
# Schema validation
# ---------------------------------------------------------------------------


def _load_product_validator(schema_dir: Path) -> Draft7Validator:
    """Load ``product.schema.json`` with ``$ref`` resolution.

    Parameters
    ----------
    schema_dir : Path
        Directory containing ``product.schema.json`` and ``_defs.json``.

    Returns
    -------
    Draft7Validator
        A prepared validator.
    """
    schema_path = schema_dir / "product.schema.json"
    schema = json.loads(schema_path.read_text(encoding="utf-8"))
    resolver = RefResolver(
        base_uri=schema_path.as_uri(),
        referrer=schema,
    )
    return Draft7Validator(schema, resolver=resolver)


def _validate_records(
    records: List[Dict[str, Any]],
    validator: Draft7Validator,
) -> bool:
    """Validate all records against the product schema.

    Errors are logged to stderr. Only the *required* product fields are checked
    — extra merged fields (price_observations, channels, etc.) are permitted.

    Parameters
    ----------
    records : list[dict]
        Parsed JSONL records.
    validator : Draft7Validator
        Pre-configured product schema validator.

    Returns
    -------
    bool
        ``True`` if all records pass schema validation.
    """
    has_errors = False
    for idx, record in enumerate(records):
        line_num = idx + 1
        for ve in validator.iter_errors(record):
            field = (
                ".".join(str(p) for p in ve.absolute_path)
                if ve.absolute_path
                else "(root)"
            )
            log_error(
                f"Line {line_num}, field '{field}': {ve.message}"
            )
            has_errors = True
    return not has_errors


# ---------------------------------------------------------------------------
# Export functions
# ---------------------------------------------------------------------------


def _write_csv(rows: List[Dict[str, Any]], path: Path) -> None:
    """Write rows to a CSV file with the standard column order.

    Parameters
    ----------
    rows : list[dict]
        Mapped output rows.
    path : Path
        Output file path (must be ``.csv``).
    """
    with open(str(path), "w", encoding="utf-8", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=OUTPUT_COLUMNS, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            # Ensure all columns present
            ordered = {col: row.get(col, "") for col in OUTPUT_COLUMNS}
            writer.writerow(ordered)


def _write_xlsx(rows: List[Dict[str, Any]], path: Path) -> None:
    """Write rows to an XLSX file with the standard column order using openpyxl.

    Parameters
    ----------
    rows : list[dict]
        Mapped output rows.
    path : Path
        Output file path (must be ``.xlsx``).
    """
    try:
        import openpyxl
        from openpyxl.utils import get_column_letter
    except ImportError:
        log_error("openpyxl is required for XLSX export — install with: pip install openpyxl")
        sys.exit(3)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Competitor Matrix"

    # Header row
    ws.append(OUTPUT_COLUMNS)

    # Data rows
    for row in rows:
        ordered = [row.get(col, "") for col in OUTPUT_COLUMNS]
        ws.append(ordered)

    # Auto-adjust column widths (capped at 40)
    for col_idx, col_name in enumerate(OUTPUT_COLUMNS, start=1):
        max_len = len(col_name)
        for row_idx in range(2, len(rows) + 2):
            cell_val = ws.cell(row=row_idx, column=col_idx).value
            if cell_val is not None:
                max_len = max(max_len, len(str(cell_val)))
        adjusted = min(max_len + 2, 40)
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted

    wb.save(str(path))


# ---------------------------------------------------------------------------
# CLI entry point
# ---------------------------------------------------------------------------


def main() -> int:
    """Parse arguments, validate, transform, and export the competitor matrix."""
    parser = argparse.ArgumentParser(
        description="Build a competitor matrix CSV/XLSX from merged product records.",
    )
    parser.add_argument(
        "input",
        type=str,
        help="Path to a .jsonl file containing merged product records",
    )
    parser.add_argument(
        "--out",
        type=str,
        required=True,
        help="Output file path (.csv or .xlsx)",
    )
    parser.add_argument(
        "--format",
        type=str,
        choices=["csv", "xlsx"],
        default=None,
        help=(
            "Output format (csv or xlsx). "
            "If omitted, inferred from --out file extension."
        ),
    )
    args = parser.parse_args()

    input_path = Path(args.input)
    output_path = Path(args.out)

    # ---- Determine format ---------------------------------------------------
    fmt = args.format
    if fmt is None:
        ext = output_path.suffix.lower()
        if ext == ".csv":
            fmt = "csv"
        elif ext == ".xlsx":
            fmt = "xlsx"
        else:
            log_error(
                f"Cannot determine format from extension '{ext}'. "
                f"Use --format csv or --format xlsx."
            )
            return 2

    if fmt not in ("csv", "xlsx"):
        log_error(f"Unsupported format: {fmt!r}. Use 'csv' or 'xlsx'.")
        return 2

    # ---- Read input ---------------------------------------------------------
    if not input_path.exists():
        log_error(f"File not found: {input_path}")
        return 2

    try:
        records = read_jsonl(input_path)
    except json.JSONDecodeError as exc:
        log_error(f"Corrupted JSONL at {input_path}: {exc}")
        return 3

    if not records:
        log_error(f"No records found in {input_path}")
        return 1

    # ---- Validate against product schema ------------------------------------
    schema_dir = _PROJECT_ROOT / "schemas"
    try:
        validator = _load_product_validator(schema_dir)
    except Exception as exc:
        log_error(f"Failed to load product.schema.json: {exc}")
        return 3

    schema_ok = _validate_records(records, validator)
    if not schema_ok:
        log_error("Schema validation failed — some records are invalid")
        # Continue with transformation but emit warnings
        # Schema errors don't necessarily mean we can't build the matrix

    # ---- Transform records --------------------------------------------------
    rows = [_extract_row(r) for r in records]

    # ---- Check required columns ---------------------------------------------
    if not _check_required_columns(rows):
        return 1

    # ---- Export -------------------------------------------------------------
    try:
        if fmt == "csv":
            _write_csv(rows, output_path)
            log_info(f"CSV exported: {output_path}")
        else:
            _write_xlsx(rows, output_path)
            log_info(f"XLSX exported: {output_path}")
    except Exception as exc:
        log_error(f"Export failed: {exc}")
        return 3

    return 0


if __name__ == "__main__":
    sys.exit(main())
